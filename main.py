import json
from pathlib import Path
from datetime import datetime  # 🔥 CORREÇÃO — Importar datetime no topo
from utils.logger import setup_logger
from utils.file_manager import OUTPUT_DIR, DOCS_PDFS_DIR
from utils.delays import pausa_estrategica
from scraper.amil_scraper import AmilBot

SCRIPT_DIR = Path(__file__).resolve().parent

# Tentar importar openpyxl, se não estiver disponível, mostrar erro
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("❌ ERRO: openpyxl não instalado!")
    print("   Instale com: pip install openpyxl")
    print("   A planilha Excel é obrigatória para o funcionamento.")
    raise


# =====================================================
# 🔹 Função para gerar/atualizar planilha Excel
# =====================================================
def gerar_planilha_simples(resultado_por_cidade: list[dict], modo_append: bool = False):
    """
    Cria ou atualiza uma planilha Excel com:
    id ; cidade ; estado ; link_pdf_publico
    
    Args:
        resultado_por_cidade: Lista de resultados a serem adicionados
        modo_append: Se True, adiciona ao arquivo existente. Se False, recria o arquivo.
    """
    planilha_dir = DOCS_PDFS_DIR / "planilhas"
    planilha_dir.mkdir(parents=True, exist_ok=True)

    caminho_xlsx = planilha_dir / "planilha_simples.xlsx"
    
    # Verificar se o arquivo já existe e se tem conteúdo
    arquivo_existe = caminho_xlsx.exists() and caminho_xlsx.stat().st_size > 0

    # Base pública do GitHub Pages
    BASE_URL = "https://rafaelsinkevicius.github.io/amil-bot/pdfs"

    # 🔥 NOVO — Ler cidades existentes do Excel para evitar duplicatas
    cidades_existentes = set()
    linhas_existentes = []
    ultimo_id = 0
    
    if arquivo_existe:
        try:
            wb_existente = load_workbook(caminho_xlsx)
            ws_existente = wb_existente.active
            
            # Ler todas as linhas (pular cabeçalho na linha 1)
            for row in ws_existente.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 3:
                    try:
                        id_atual = int(row[0]) if row[0] else 0
                        cidade = str(row[1]) if row[1] else ""
                        uf = str(row[2]) if row[2] else ""
                        
                        if cidade and uf:
                            cidade_uf = f"{cidade}-{uf}"
                            cidades_existentes.add(cidade_uf)
                            linhas_existentes.append([id_atual, cidade, uf, row[3] if len(row) >= 4 else ""])
                            
                            if id_atual > ultimo_id:
                                ultimo_id = id_atual
                    except:
                        continue
        except Exception as e:
            print(f"⚠️  Erro ao ler planilha Excel existente: {e}")

    # Filtrar cidades que já existem
    novas_cidades = []
    for item in resultado_por_cidade:
        cidade_uf = f"{item['cidade']}-{item['uf']}"
        if cidade_uf not in cidades_existentes:
            novas_cidades.append(item)
        else:
            print(f"⏭️  Cidade {cidade_uf} já existe na planilha, pulando...")

    if not novas_cidades:
        print("📋 Nenhuma cidade nova para adicionar à planilha.")
        return

    # ========== GERAR EXCEL (com links clicáveis) ==========
    try:
        # 🔥 CORREÇÃO: sempre carregar se arquivo existe, independente de modo_append
        if arquivo_existe:
            # Carregar workbook existente
            wb = load_workbook(caminho_xlsx)
            ws = wb.active
            
            # 🔥 NOVO — Garantir que cabeçalho existe
            if ws.max_row == 0 or ws.cell(row=1, column=1).value != "ID":
                headers = ["ID", "Cidade", "Estado", "Link PDF"]
                ws.insert_rows(1)
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col)
                    cell.value = header
                    cell.font = Font(bold=True, size=12)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            # Criar novo workbook apenas se arquivo não existe
            wb = Workbook()
            ws = wb.active
            ws.title = "Planilha PDFs"
            
            # Cabeçalhos
            headers = ["ID", "Cidade", "Estado", "Link PDF"]
            ws.append(headers)
            
            # Formatar cabeçalho
            header_font = Font(bold=True, size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.alignment = header_alignment
        
        # Adicionar novas cidades
        for item in novas_cidades:
            ultimo_id += 1
            cidade = item["cidade"]
            uf = item["uf"]

            nome_arquivo = f"{cidade}-{uf}".replace(" ", "_")
            pdf_name = f"{nome_arquivo}.pdf"
            link_publico = f"{BASE_URL}/{uf}/{pdf_name}"
            
            # Adicionar linha com link clicável
            row = [ultimo_id, cidade, uf, link_publico]
            ws.append(row)
            
            # Tornar o link clicável
            link_cell = ws.cell(row=ws.max_row, column=4)
            link_cell.hyperlink = link_publico
            link_cell.font = Font(color="0000FF", underline="single")
            link_cell.value = link_publico
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 70
        
        # Congelar primeira linha
        ws.freeze_panes = "A2"
        
        wb.save(caminho_xlsx)
        print(f"📊 Planilha Excel {'atualizada' if arquivo_existe else 'gerada'} em: {caminho_xlsx}")
        
    except Exception as e:
        print(f"❌ Erro ao gerar Excel: {e}")
        import traceback
        traceback.print_exc()


# =====================================================
# 🔹 Funções para salvar/carregar progresso
# =====================================================
def salvar_progresso(uf: str, cidade: str) -> None:
    """Salva o progresso atual (última cidade processada)."""
    progresso_path = OUTPUT_DIR / "progresso.json"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    progresso = {
        "uf": uf,
        "cidade": cidade,
        "timestamp": datetime.now().isoformat()
    }
    
    with open(progresso_path, "w", encoding="utf-8") as f:
        json.dump(progresso, f, ensure_ascii=False, indent=2)

def carregar_progresso() -> dict | None:
    """Carrega o último progresso salvo."""
    progresso_path = OUTPUT_DIR / "progresso.json"
    
    if not progresso_path.exists():
        return None
    
    try:
        with open(progresso_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return None

def limpar_progresso() -> None:
    """Limpa o arquivo de progresso."""
    progresso_path = OUTPUT_DIR / "progresso.json"
    if progresso_path.exists():
        progresso_path.unlink()

# =====================================================
# Carregar arquivo JSON das cidades
# =====================================================
def carregar_mapa_estados() -> dict:
    caminho_json = SCRIPT_DIR / "estados_cidades_amil.json"
    with open(caminho_json, encoding="utf-8") as f:
        return json.load(f)


# =====================================================
# Logs finais (salva em output/ apenas para logs)
# =====================================================
def salvar_logs_finais(resultado_por_cidade, cidades_com_erro) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    caminho_json_erros = OUTPUT_DIR / "cidades_com_erro.json"
    with open(caminho_json_erros, "w", encoding="utf-8") as f:
        json.dump(cidades_com_erro, f, ensure_ascii=False, indent=2)

    caminho_log = OUTPUT_DIR / "log_execucao.txt"
    total_ok = len(resultado_por_cidade)
    total_erro = sum(len(cidades) for cidades in cidades_com_erro.values())

    with open(caminho_log, "w", encoding="utf-8") as log:
        log.write("📊 RESUMO FINAL\n")
        log.write(f"✅ Cidades concluídas com sucesso: {total_ok}\n")
        log.write(f"❌ Cidades com erro: {total_erro}\n\n")
        log.write("📍 Prestadores por cidade:\n")
        for item in sorted(resultado_por_cidade, key=lambda x: (x["uf"], x["cidade"])):
            log.write(
                f"- {item['cidade']}/{item['uf']}: {item['prestadores']} prestadores\n"
            )


# =====================================================
# MAIN
# =====================================================
def main() -> None:
    """Executa o bot normalmente."""
    executar_bot_com_callbacks(None, None, None)

def executar_bot_com_callbacks(callback_progresso=None, callback_log=None, stop_flag=None, continuar_progresso: bool = True):
    """
    Executa o bot com callbacks para interface web.
    
    Args:
        callback_progresso: Função(uf, cidade, total, atual) chamada a cada cidade
        callback_log: Função(mensagem) chamada para logs
        stop_flag: threading.Event para parar execução
        continuar_progresso: Se True, continua de onde parou. Se False, começa do zero.
    """
    
    logger = setup_logger("amil_bot", OUTPUT_DIR / "amil_bot.log")
    mapa = carregar_mapa_estados()

    resultado_por_cidade_global = []
    cidades_com_erro_global = {}
    contador_cidades = 0
    
    # 🔥 NOVO — Carregar progresso anterior
    progresso_anterior = None
    pular_ate_progresso = False
    
    if continuar_progresso:
        progresso_anterior = carregar_progresso()
        if progresso_anterior:
            if callback_log:
                callback_log(f"📌 Continuando de: {progresso_anterior['cidade']}-{progresso_anterior['uf']}")
            pular_ate_progresso = True
    
    # Calcular total de cidades
    total_cidades = sum(len(cidades) for cidades in mapa.values())
    
    if callback_log:
        callback_log(f"Total de cidades a processar: {total_cidades}")
    
    # Flag para controlar se já criou o cabeçalho da planilha
    primeira_vez = True

    try:
        for uf, cidades in mapa.items():
            if stop_flag and stop_flag.is_set():
                if callback_log:
                    callback_log("⛔ Execução interrompida pelo usuário")
                break
                
            if callback_log:
                callback_log(f"Iniciando UF {uf} ({len(cidades)} cidades)")
            
            logger.info(f"====== Iniciando UF {uf} ({len(cidades)} cidades) ======")
            
            with AmilBot(uf, pasta_base=DOCS_PDFS_DIR, logger=logger, stop_flag=stop_flag) as bot:
                for cidade in cidades:
                    if stop_flag and stop_flag.is_set():
                        if callback_log:
                            callback_log("⛔ Execução interrompida pelo usuário")
                        break
                    
                    # 🔥 NOVO — Pular cidades até chegar no progresso anterior
                    if pular_ate_progresso:
                        if progresso_anterior:
                            if uf == progresso_anterior["uf"] and cidade == progresso_anterior["cidade"]:
                                if callback_log:
                                    callback_log(f"✅ Continuando após {cidade}-{uf} (pulando cidade já processada)")
                                pular_ate_progresso = False
                                # 🔥 CORREÇÃO — Atualizar contador e callback ao pular
                                contador_cidades += 1
                                if callback_progresso:
                                    callback_progresso(uf, cidade, total_cidades, contador_cidades)
                                # Pular a cidade do progresso e começar na próxima
                                continue
                            else:
                                # 🔥 CORREÇÃO — Atualizar contador e callback ao pular cidades anteriores
                                contador_cidades += 1
                                if callback_progresso:
                                    callback_progresso(uf, cidade, total_cidades, contador_cidades)
                                if callback_log:
                                    callback_log(f"⏭️  Pulando {cidade}-{uf} (já processada)")
                                # Ainda não chegou no progresso, pular
                                continue
                        else:
                            pular_ate_progresso = False
                    
                    # Callback de progresso
                    if callback_progresso:
                        callback_progresso(uf, cidade, total_cidades, contador_cidades)
                    
                    # 🔥 NOVO — Salvar progresso ANTES de processar (para caso trave)
                    salvar_progresso(uf, cidade)
                    
                    # 🔥 NOVO — Timeout máximo para processar cidade (5 minutos)
                    import time as time_module
                    inicio_processamento = time_module.time()
                    timeout_maximo_cidade = 300  # 5 minutos
                    
                    try:
                        bot.processar_cidade(cidade)
                    except Exception as e:
                        # Verificar se foi timeout
                        if time_module.time() - inicio_processamento > timeout_maximo_cidade:
                            if callback_log:
                                callback_log(f"⏱️ Timeout ao processar {cidade}-{uf} (mais de 5 minutos)")
                        raise

                    # coleta resultados
                    resultado_por_cidade_global.extend(bot.resultado_por_cidade)
                    for k, v in bot.cidades_com_erro.items():
                        cidades_com_erro_global.setdefault(k, []).extend(v)

                    # salvar planilha incrementalmente após cada cidade
                    if bot.resultado_por_cidade:
                        gerar_planilha_simples(
                            bot.resultado_por_cidade, 
                            modo_append=not primeira_vez
                        )
                        primeira_vez = False
                        
                        if callback_log:
                            for item in bot.resultado_por_cidade:
                                if item.get("prestadores", 0) > 0:
                                    callback_log(f"✅ {item['cidade']}-{item['uf']}: {item['prestadores']} prestadores encontrados")
                                else:
                                    callback_log(f"⚠️ {item['cidade']}-{item['uf']}: PDF vazio gerado (sem especialidade)")
                    
                    # 🔥 CORREÇÃO — Salvar progresso sempre, mesmo sem resultados
                    salvar_progresso(uf, cidade)

                    # limpa buffers do bot
                    bot.resultado_por_cidade.clear()
                    bot.cidades_com_erro.clear()

                    contador_cidades += 1
                    pausa_estrategica(contador_cidades)

    except KeyboardInterrupt:
        logger.warning("⛔ Execução interrompida manualmente. Gerando logs parciais...")
        if callback_log:
            callback_log("⛔ Execução interrompida manualmente")

    # salva logs normais
    salvar_logs_finais(resultado_por_cidade_global, cidades_com_erro_global)

    logger.info("✅ Execução finalizada.")
    if callback_log:
        callback_log("✅ Execução finalizada")
    
    # 🔥 NOVO — Limpar progresso quando terminar tudo
    if not stop_flag or not stop_flag.is_set():
        limpar_progresso()
        if callback_log:
            callback_log("📌 Progresso limpo - todas as cidades foram processadas")


if __name__ == "__main__":
    main()